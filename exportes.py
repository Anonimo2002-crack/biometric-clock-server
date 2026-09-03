"""PDF y Excel de los reportes del tablero."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from reportes import INSTITUCION, JORNADA

_FONT_NAME = "Helvetica"
_FONT_BOLD = "Helvetica-Bold"


def _registrar_fuente() -> None:
    global _FONT_NAME, _FONT_BOLD
    candidatos = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
        Path("/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf"),
    ]
    bold_map = {
        "arial.ttf": Path("C:/Windows/Fonts/arialbd.ttf"),
        "calibri.ttf": Path("C:/Windows/Fonts/calibrib.ttf"),
        "DejaVuSans.ttf": Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        "LiberationSans-Regular.ttf": Path(
            "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"
        ),
    }
    for ruta in candidatos:
        if not ruta.is_file():
            continue
        pdfmetrics.registerFont(TTFont("Asistencia", str(ruta)))
        _FONT_NAME = "Asistencia"
        bold = bold_map.get(ruta.name)
        if bold and bold.is_file():
            pdfmetrics.registerFont(TTFont("Asistencia-Bold", str(bold)))
            _FONT_BOLD = "Asistencia-Bold"
        else:
            _FONT_BOLD = "Asistencia"
        return


_registrar_fuente()

_PINO = colors.HexColor("#1b2a24")
_AMBAR = colors.HexColor("#e8a317")
_PAPEL = colors.HexColor("#fff8ee")
_LINEA = colors.HexColor("#e4d4be")
_EXCEL_HEADER = PatternFill("solid", fgColor="1B2A24")
_EXCEL_GOLD = PatternFill("solid", fgColor="F3C56B")


def _estilos_pdf() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "kicker": ParagraphStyle(
            "kicker",
            parent=base["Normal"],
            fontName=_FONT_BOLD,
            fontSize=9,
            textColor=_AMBAR,
            spaceAfter=2,
        ),
        "titulo": ParagraphStyle(
            "titulo",
            parent=base["Heading1"],
            fontName=_FONT_BOLD,
            fontSize=16,
            textColor=_PINO,
            spaceAfter=4,
        ),
        "meta": ParagraphStyle(
            "meta",
            parent=base["Normal"],
            fontName=_FONT_NAME,
            fontSize=9,
            textColor=colors.HexColor("#6b5f52"),
            spaceAfter=10,
        ),
    }


def armar_pdf(titulo: str, subtitulo: str, encabezados: list[str], filas: list[list[Any]]) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter,
        leftMargin=1.6 * cm,
        rightMargin=1.6 * cm,
        topMargin=1.4 * cm,
        bottomMargin=1.4 * cm,
        title=titulo,
        author=INSTITUCION,
    )
    estilos = _estilos_pdf()
    story: list[Any] = [
        Paragraph(INSTITUCION, estilos["kicker"]),
        Paragraph(titulo, estilos["titulo"]),
        Paragraph(f"{JORNADA} · {subtitulo}", estilos["meta"]),
        Spacer(1, 6),
    ]
    datos = [encabezados, *[[("" if cell is None else str(cell)) for cell in fila] for fila in filas]]
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(
        TableStyle(
            [
                ("FONTNAME", (0, 0), (-1, 0), _FONT_BOLD),
                ("FONTNAME", (0, 1), (-1, -1), _FONT_NAME),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), _PINO),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, -1), _PAPEL),
                ("GRID", (0, 0), (-1, -1), 0.3, _LINEA),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(tabla)
    if not filas:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Sin registros para esta fecha.", estilos["meta"]))
    doc.build(story)
    return buffer.getvalue()


def armar_excel(titulo: str, subtitulo: str, encabezados: list[str], filas: list[list[Any]]) -> bytes:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Reporte"
    hoja["A1"] = INSTITUCION
    hoja["A1"].font = Font(bold=True, color="1B2A24", size=14)
    hoja["A2"] = titulo
    hoja["A2"].font = Font(bold=True, size=12)
    hoja["A3"] = f"{JORNADA} · {subtitulo}"
    hoja["A3"].font = Font(italic=True, color="6B5F52")

    for col, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=5, column=col, value=texto)
        celda.font = Font(bold=True, color="FFFFFF", size=10)
        celda.fill = _EXCEL_HEADER
        celda.alignment = Alignment(horizontal="left")

    for i, fila in enumerate(filas, start=6):
        for col, valor in enumerate(fila, start=1):
            hoja.cell(row=i, column=col, value="" if valor is None else valor)

    for col in range(1, len(encabezados) + 1):
        letra = get_column_letter(col)
        largo = max(
            len(str(encabezados[col - 1])),
            *(len(str(fila[col - 1] or "")) for fila in filas),
            12,
        )
        hoja.column_dimensions[letra].width = min(largo + 4, 42)

    buffer = BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def dashboard_filas(dto: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    encabezados = ["Grado", "Matriculados", "Presentes", "Tarde", "Ausentes", "%"]
    filas = [
        [
            item["grado"],
            item["matriculados"],
            item["presentes"],
            item["tardes"],
            item["ausentes"],
            item["porcentaje"],
        ]
        for item in dto.get("porGrado") or []
    ]
    return encabezados, filas


def asistencia_filas(dto: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    encabezados = ["Alumno", "Grado", "Hora de marca", "Estado"]
    filas = [
        [item["nombre"], item["grado"], item.get("horaMarca") or "Sin marca", item["estado"]]
        for item in dto.get("alumnos") or []
    ]
    return encabezados, filas


def ausencias_filas(dto: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    encabezados = ["Alumno", "Grado", "Marca", "Estado"]
    filas = [
        [item["nombre"], item["grado"], item.get("horaMarca") or "Sin marca", item["estado"]]
        for item in dto.get("alumnos") or []
    ]
    return encabezados, filas


def maestros_filas(dto: dict[str, Any]) -> tuple[list[str], list[list[Any]]]:
    encabezados = ["Maestro", "Cargo", "Entrada", "Salida", "Estado"]
    filas = [
        [
            item["nombre"],
            item.get("cargo") or "",
            item.get("horaEntrada") or "—",
            item.get("horaSalida") or "—",
            item["estado"],
        ]
        for item in dto.get("maestros") or []
    ]
    return encabezados, filas
